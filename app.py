import os
from flask import Flask, render_template, request, redirect, url_for, flash, send_file
from flask_login import LoginManager, login_user, login_required, logout_user, current_user
from werkzeug.utils import secure_filename
import qrcode
import csv
import io
from datetime import datetime
from extensions import app, db
from models import Сотрудник, Документ, ПрочтениеДокумента, Отдел, Организация
import asyncio
from telegram.ext import Application
from config import Config
from bot import notify_new_document
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Инициализация Flask-Login
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

@login_manager.user_loader
def load_user(user_id):
    return Сотрудник.query.get(int(user_id))

# Глобальная переменная для бота
bot_app = None

async def setup_bot():
    """Инициализация бота при запуске приложения"""
    global bot_app
    if not bot_app:
        bot_app = Application.builder().token(Config.TELEGRAM_BOT_TOKEN).build()
        await bot_app.initialize()
        await bot_app.start()
    return bot_app

@app.route('/')
@login_required
def index():
    return redirect(url_for('documents'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')
        
        user = Сотрудник.query.filter_by(email=email).first()
        if user and check_password_hash(user.пароль, password):
            login_user(user)
            return redirect(url_for('documents'))
        
        flash('Неверный email или пароль')
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/documents')
@login_required
def documents():
    # Получаем параметры фильтрации
    отдел_id = request.args.get('отдел_id', type=int)
    отдел_сотрудников_id = request.args.get('отдел_сотрудников_id', type=int)
    статус_telegram = request.args.get('статус_telegram')
    
    # Базовый запрос для документов
    query = Документ.query
    
    # Если пользователь не администратор, показываем только документы его отдела
    if current_user.роль != 'администратор':
        query = query.filter(Документ.отдел_id == current_user.отдел_id)
    elif отдел_id:  # Для администратора применяем фильтр по отделу
        query = query.filter(Документ.отдел_id == отдел_id)
    
    документы = query.all()
    
    # Для каждого документа подсчитываем статистику ознакомления
    for док in документы:
        total_employees = Сотрудник.query.filter_by(отдел_id=док.отдел_id).count()
        read_count = ПрочтениеДокумента.query.filter_by(документ_id=док.id).count()
        док.progress = int((read_count / total_employees * 100) if total_employees > 0 else 0)
        док.read_count = read_count
        док.total_count = total_employees
    
    # Получаем список сотрудников с фильтрацией
    сотрудники_query = Сотрудник.query
    
    if отдел_сотрудников_id:
        сотрудники_query = сотрудники_query.filter_by(отдел_id=отдел_сотрудников_id)
    
    if статус_telegram:
        if статус_telegram == 'connected':
            сотрудники_query = сотрудники_query.filter(Сотрудник.статус_регистрации == True)
        elif статус_telegram == 'not_connected':
            сотрудники_query = сотрудники_query.filter(
                (Сотрудник.статус_регистрации == False) | (Сотрудник.telegram_id == None)
            )
    
    сотрудники = сотрудники_query.all()
    
    # Получаем список всех отделов для фильтра
    отделы = Отдел.query.all()
    
    return render_template('documents.html', 
                         документы=документы,
                         сотрудники=сотрудники,
                         отделы=отделы,
                         current_filter={
                             'отдел_id': отдел_id,
                             'отдел_сотрудников_id': отдел_сотрудников_id,
                             'статус_telegram': статус_telegram
                         })

@app.route('/upload', methods=['GET'])
@login_required
def upload_document_form():
    if current_user.роль != 'администратор':
        flash('Недостаточно прав для загрузки документов')
        return redirect(url_for('documents'))
    
    # Получаем список всех отделов
    отделы = Отдел.query.all()
    return render_template('upload.html', отделы=отделы)

@app.route('/upload', methods=['POST'])
@login_required
def upload_document():
    if current_user.роль != 'администратор':
        flash('Недостаточно прав для загрузки документов')
        return redirect(url_for('documents'))
        
    try:
        # Получение данных формы
        название = request.form.get('title')
        тип_документа = request.form.get('type')
        срок_ознакомления = datetime.strptime(request.form.get('deadline'), '%Y-%m-%d')
        file = request.files.get('file')
        отдел_id = request.form.get('department')  # Получаем ID отдела из формы
        
        if not all([название, тип_документа, срок_ознакомления, file]):
            flash('Пожалуйста, заполните все поля')
            return redirect(url_for('upload_document_form'))
            
        # Создание директории для загрузок, если её нет
        upload_dir = app.config['UPLOAD_FOLDER']
        if not os.path.exists(upload_dir):
            os.makedirs(upload_dir, exist_ok=True)
            
        # Сохранение файла
        filename = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{secure_filename(file.filename)}"
        file_path = os.path.join(upload_dir, filename)
        file.save(file_path)

        # Создание директории для QR-кодов
        qr_dir = os.path.join('static', 'qr_codes')
        os.makedirs(qr_dir, exist_ok=True)

        if отдел_id == 'all':
            # Если выбраны все отделы, создаем документ для каждого отдела
            отделы = Отдел.query.all()
            for отдел in отделы:
                документ = Документ(
                    название=название,
                    тип_документа=тип_документа,
                    срок_ознакомления=срок_ознакомления,
                    путь_к_файлу=filename,
                    дата_создания=datetime.now(),
                    отдел_id=отдел.id
                )
                db.session.add(документ)
                db.session.flush()  # Получаем ID документа
                
                # Генерация QR-кода
                qr = qrcode.QRCode(version=1, box_size=10, border=5)
                qr.add_data(str(документ.id))
                qr.make(fit=True)
                
                # Сохранение QR-кода
                qr_filename = f'document_{документ.id}.png'
                qr_path = os.path.join(qr_dir, qr_filename)
                qr.make_image(fill_color="black", back_color="white").save(qr_path)
                
                # Обновляем запись в базе данных с путем к QR-коду
                документ.qr_код = qr_filename
                
                # После создания документа отправляем уведомление
                asyncio.run(notify_new_document(документ.id))
        else:
            # Создание записи в базе данных для одного отдела
            документ = Документ(
                название=название,
                тип_документа=тип_документа,
                срок_ознакомления=срок_ознакомления,
                путь_к_файлу=filename,
                дата_создания=datetime.now(),
                отдел_id=int(отдел_id)
            )
            db.session.add(документ)
            db.session.flush()  # Получаем ID документа
            
            # Генерация QR-кода
            qr = qrcode.QRCode(version=1, box_size=10, border=5)
            qr.add_data(str(документ.id))
            qr.make(fit=True)
            
            # Сохранение QR-кода
            qr_filename = f'document_{документ.id}.png'
            qr_path = os.path.join(qr_dir, qr_filename)
            qr.make_image(fill_color="black", back_color="white").save(qr_path)
            
            # Обновляем запись в базе данных с путем к QR-коду
            документ.qr_код = qr_filename
            
            # После создания документа отправляем уведомление
            asyncio.run(notify_new_document(документ.id))
            
        db.session.commit()
        flash('Документ успешно загружен')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Ошибка при загрузке документа: {str(e)}')
        return redirect(url_for('upload_document_form'))
        
    return redirect(url_for('documents'))

@app.route('/get_qr_code/<int:document_id>')
@login_required
def get_qr_code(document_id):
    try:
        документ = Документ.query.get_or_404(document_id)
        if not документ.qr_код:
            # Если QR-код еще не создан, создаем его
            qr = qrcode.QRCode(version=1, box_size=10, border=5)
            qr.add_data(str(документ.id))
            qr.make(fit=True)
            
            # Создаем директорию для QR-кодов, если её нет
            qr_dir = os.path.join('static', 'qr_codes')
            os.makedirs(qr_dir, exist_ok=True)
            
            # Сохраняем QR-код
            qr_filename = f'document_{документ.id}.png'
            qr_path = os.path.join(qr_dir, qr_filename)
            qr.make_image(fill_color="black", back_color="white").save(qr_path)
            
            # Обновляем запись в базе данных
            документ.qr_код = qr_filename
            db.session.commit()
        
        return send_file(
            os.path.join('static', 'qr_codes', документ.qr_код),
            mimetype='image/png'
        )
    except Exception as e:
        flash(f'Ошибка при генерации QR-кода: {str(e)}')
        return redirect(url_for('documents'))

@app.route('/download/<int:document_id>')
@login_required
def download_document(document_id):
    документ = Документ.query.get_or_404(document_id)
    if документ.отдел_id != current_user.отдел_id:
        return 'Доступ запрещен', 403
    
    file_path = os.path.join(app.root_path, 'uploads', документ.путь_к_файлу)
    return send_file(file_path, 
                    as_attachment=True,
                    download_name=документ.путь_к_файлу.split('_', 2)[-1])

@app.route('/statistics')
@login_required
def statistics():
    if current_user.роль not in ['администратор', 'руководитель']:
        flash('Недостаточно прав для просмотра статистики')
        return redirect(url_for('documents'))

    # Получаем параметры фильтрации
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    document_type = request.args.get('document_type')
    department_id = request.args.get('department_id')
    
    # Базовый запрос
    query = db.session.query(
        Документ,
        Отдел.название.label('отдел_название'),
        db.func.count(Сотрудник.id).label('всего_сотрудников'),
        db.func.count(ПрочтениеДокумента.id).label('ознакомились')
    ).join(
        Отдел, Документ.отдел_id == Отдел.id
    ).outerjoin(
        Сотрудник, Сотрудник.отдел_id == Отдел.id
    ).outerjoin(
        ПрочтениеДокумента, 
        db.and_(
            ПрочтениеДокумента.документ_id == Документ.id,
            ПрочтениеДокумента.сотрудник_id == Сотрудник.id
        )
    )
    
    # Применяем фильтры
    if start_date:
        query = query.filter(Документ.дата_создания >= datetime.strptime(start_date, '%Y-%m-%d'))
    if end_date:
        query = query.filter(Документ.дата_создания <= datetime.strptime(end_date, '%Y-%m-%d'))
    if document_type:
        query = query.filter(Документ.тип_документа == document_type)
    if department_id:
        query = query.filter(Отдел.id == department_id)
    
    # Группировка
    stats = query.group_by(Документ.id, Отдел.id).all()
    
    # Преобразуем результаты
    статистика = []
    for док, отдел_название, всего, ознакомились in stats:
        статистика.append({
            'документ': док,
            'отдел': отдел_название,
            'всего_сотрудников': всего,
            'ознакомились': ознакомились,
            'процент': round((ознакомились / всего * 100) if всего > 0 else 0, 2)
        })
    
    # Получаем данные для фильтров
    отделы = Отдел.query.all()
    типы_документов = db.session.query(Документ.тип_документа).distinct().all()
    
    return render_template(
        'statistics.html',
        статистика=статистика,
        отделы=отделы,
        типы_документов=типы_документов,
        current_filters={
            'start_date': start_date,
            'end_date': end_date,
            'document_type': document_type,
            'department_id': department_id
        }
    )

@app.route('/export_statistics')
@login_required
def export_statistics():
    """Экспорт статистики в Excel"""
    if current_user.роль != 'администратор':
        flash('Недостаточно прав для экспорта статистики')
        return redirect(url_for('statistics'))

    try:
        # Получаем параметры фильтрации
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        document_type = request.args.get('document_type')
        department_id = request.args.get('department_id')
        if department_id:
            department_id = int(department_id)

        # Базовый запрос
        query = db.session.query(
            Документ,
            Отдел.название.label('отдел'),
            db.func.count(Сотрудник.id).label('всего_сотрудников'),
            db.func.count(ПрочтениеДокумента.id).label('прочитали')
        ).join(
            Отдел, Документ.отдел_id == Отдел.id
        ).outerjoin(
            Сотрудник, Сотрудник.отдел_id == Отдел.id
        ).outerjoin(
            ПрочтениеДокумента,
            db.and_(
                ПрочтениеДокумента.документ_id == Документ.id,
                ПрочтениеДокумента.сотрудник_id == Сотрудник.id
            )
        )

        # Применяем фильтры
        if start_date:
            query = query.filter(Документ.дата_создания >= datetime.strptime(start_date, '%Y-%m-%d'))
        if end_date:
            query = query.filter(Документ.дата_создания <= datetime.strptime(end_date, '%Y-%m-%d'))
        if document_type:
            query = query.filter(Документ.тип_документа == document_type)
        if department_id:
            query = query.filter(Отдел.id == department_id)

        # Группировка
        stats = query.group_by(Документ.id, Отдел.id).all()

        # Создаем новую книгу Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Статистика"

        # Стили для заголовков
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Заголовки
        headers = [
            "Название документа", "Тип документа", "Отдел",
            "Дата создания", "Срок ознакомления",
            "Всего сотрудников", "Прочитали", "Процент ознакомления"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # Заполняем данные
        for row, (док, отдел, всего, прочитали) in enumerate(stats, 2):
            ws.cell(row=row, column=1, value=док.название)
            ws.cell(row=row, column=2, value=док.тип_документа)
            ws.cell(row=row, column=3, value=отдел)
            ws.cell(row=row, column=4, value=док.дата_создания.strftime('%d.%m.%Y'))
            ws.cell(row=row, column=5, value=док.срок_ознакомления.strftime('%d.%m.%Y'))
            ws.cell(row=row, column=6, value=всего)
            ws.cell(row=row, column=7, value=прочитали)
            процент = round((прочитали / всего * 100) if всего > 0 else 0, 2)
            ws.cell(row=row, column=8, value=f"{процент}%")

        # Автоматическая ширина столбцов
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        # Сохраняем в буфер
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='statistics.xlsx'
        )

    except Exception as e:
        flash(f'Ошибка при экспорте статистики: {str(e)}')
        return redirect(url_for('statistics'))

@app.route('/add_department', methods=['GET', 'POST'])
@login_required
def add_department():
    if current_user.роль != 'администратор':
        flash('Недостаточно прав для добавления отделов')
        return redirect(url_for('departments'))
        
    if request.method == 'POST':
        название = request.form.get('name')
        if not название:
            flash('Название отдела не может быть пустым')
            return redirect(url_for('add_department'))
            
        организация = Организация.query.first()
        if not организация:
            # Если организации нет, создаем её
            организация = Организация(название='Основная организация')
            db.session.add(организация)
            db.session.flush()
        
        отдел = Отдел(название=название, организация_id=организация.id)
        db.session.add(отдел)
        db.session.commit()
        
        flash('Отдел успешно добавлен')
        return redirect(url_for('departments'))
        
    return render_template('add_department.html')

@app.route('/add_employee', methods=['GET', 'POST'])
@login_required
def add_employee():
    if current_user.роль != 'администратор':
        flash('Недостаточно прав для добавления сотрудников')
        return redirect(url_for('users'))
        
    if request.method == 'POST':
        new_employee = Сотрудник(
            фамилия=request.form['surname'],
            имя=request.form['name'],
            отчество=request.form['patronymic'],
            табельный_номер=request.form['employee_id'],
            должность=request.form['position'],
            email=request.form['email'],
            пароль=generate_password_hash(request.form['password']),
            отдел_id=request.form['department'],
            роль=request.form['role']
        )
        
        try:
            db.session.add(new_employee)
            db.session.commit()
            flash('Сотрудник успешно добавлен')
            return redirect(url_for('users'))
        except Exception as e:
            db.session.rollback()
            flash(f'Ошибка при добавлении сотрудника: {str(e)}')
            
    отделы = Отдел.query.all()
    return render_template('add_employee.html', отделы=отделы)

@app.route('/delete_document/<int:id>', methods=['POST'])
@login_required
def delete_document(id):
    if current_user.роль != 'администратор':
        flash('Недостаточно прав для удаления документов')
        return redirect(url_for('documents'))

    документ = Документ.query.get_or_404(id)
    
    # Удаляем связанные записи
    ПрочтениеДокумента.query.filter_by(документ_id=id).delete()
    
    # Удаляем файл документа
    if документ.путь_к_файлу:
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], документ.путь_к_файлу)
        if os.path.exists(file_path):
            os.remove(file_path)
    
    # Удаляем QR-код
    if документ.qr_код:
        qr_path = os.path.join('static', 'qr_codes', документ.qr_код)
        if os.path.exists(qr_path):
            os.remove(qr_path)
    
    db.session.delete(документ)
    db.session.commit()
    
    flash('Документ успешно удален')
    return redirect(url_for('documents'))

@app.route('/edit_department/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_department(id):
    if current_user.роль != 'администратор':
        flash('Недостаточно прав для редактирования отделов')
        return redirect(url_for('departments'))

    отдел = Отдел.query.get_or_404(id)
    
    if request.method == 'POST':
        отдел.название = request.form['name']
        db.session.commit()
        flash('Отдел успешно обновлен')
        return redirect(url_for('departments'))
        
    return render_template('edit_department.html', отдел=отдел)

@app.route('/departments')
@login_required
def departments():
    if current_user.роль != 'администратор':
        flash('Недостаточно прав для просмотра отделов')
        return redirect(url_for('documents'))
        
    отделы = Отдел.query.all()
    return render_template('departments.html', отделы=отделы)

@app.route('/edit_position/<int:user_id>', methods=['GET', 'POST'])
@login_required
def edit_position(user_id):
    if current_user.роль != 'администратор':
        flash('Недостаточно прав для редактирования данных сотрудника')
        return redirect(url_for('users'))

    сотрудник = Сотрудник.query.get_or_404(user_id)
    отделы = Отдел.query.all()
    
    if request.method == 'POST':
        try:
            # Обновляем основные данные
            сотрудник.фамилия = request.form['фамилия']
            сотрудник.имя = request.form['имя']
            сотрудник.отчество = request.form['отчество']
            сотрудник.табельный_номер = request.form['табельный_номер']
            сотрудник.email = request.form['email']
            сотрудник.должность = request.form['должность']
            сотрудник.рабочий_телефон = request.form['рабочий_телефон']
            сотрудник.отдел_id = int(request.form['отдел_id'])
            сотрудник.роль = request.form['роль']
            
            db.session.commit()
            flash('Данные сотрудника успешно обновлены')
            return redirect(url_for('users'))
        except Exception as e:
            db.session.rollback()
            flash(f'Ошибка при обновлении данных: {str(e)}')
            return render_template('edit_position.html', сотрудник=сотрудник, отделы=отделы)
        
    return render_template('edit_position.html', сотрудник=сотрудник, отделы=отделы)

@app.route('/users')
@login_required
def users():
    if current_user.роль != 'администратор':
        flash('Недостаточно прав для просмотра пользователей')
        return redirect(url_for('documents'))
        
    сотрудники = Сотрудник.query.all()
    return render_template('users.html', сотрудники=сотрудники)

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(debug=True)
